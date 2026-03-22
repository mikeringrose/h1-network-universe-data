"""Multi-sheet XLSX ingestion for CMS HSD Reference File."""

from pathlib import Path

import openpyxl
import polars as pl

from pipeline.config import get_settings


def read_all_sheets(path: str | Path) -> dict[str, pl.DataFrame]:
    """Read every sheet from the workbook as raw row data (no header promotion).
    Path resolved relative to DATA_DIR if not absolute. Preserves full column width
    despite merged cells in the first row."""
    settings = get_settings()
    resolved = Path(path) if Path(path).is_absolute() else Path(settings.data_dir) / path
    if not resolved.exists():
        raise FileNotFoundError(resolved)

    wb = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
    result: dict[str, pl.DataFrame] = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result[name] = pl.DataFrame()
                continue
            n_cols = max(len(r) for r in rows)
            padded = [tuple(r) + (None,) * (n_cols - len(r)) for r in rows]
            result[name] = pl.DataFrame(padded, orient="row")
    finally:
        wb.close()
    return result
